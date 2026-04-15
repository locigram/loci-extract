"""Tests for the XLSX extractor path (loci_extract/xlsx.py).

Uses openpyxl to build small in-memory workbooks; no LLM calls. The live
QB-XLSX end-to-end runs are in test_real_fixtures.py, guarded by
``LOCI_EXTRACT_LIVE_LLM=1``.
"""

from __future__ import annotations

import pytest

pytest.importorskip("openpyxl")

from loci_extract.detector import detect_financial_document_type
from loci_extract.xlsx import extract_xlsx_text
from tests.fixtures import FIXTURE_REGISTRY, skip_if_no_fixture


def _make_xlsx(tmp_path, sheets: dict[str, list[list]]):
    """Build a multi-sheet .xlsx from in-memory rows."""
    import openpyxl
    wb = openpyxl.Workbook()
    # Remove the default sheet that Workbook() creates
    wb.remove(wb.active)
    for name, rows in sheets.items():
        ws = wb.create_sheet(name)
        for row in rows:
            ws.append(row)
    path = tmp_path / "test.xlsx"
    wb.save(str(path))
    return path


def test_extract_skips_qb_tips_sheet(tmp_path):
    path = _make_xlsx(tmp_path, {
        "QuickBooks Desktop Export Tips": [["Boilerplate tip"]],
        "Sheet1": [["Real", "Data"], ["Row 1", 100]],
    })
    text = extract_xlsx_text(path)
    assert "Boilerplate" not in text
    assert "Real" in text
    assert "--- SHEET: Sheet1 ---" in text


def test_extract_preserves_indentation(tmp_path):
    # Column index 0 = depth 0, column 2 = depth 2 (4-space indent)
    path = _make_xlsx(tmp_path, {
        "Sheet1": [
            ["Top Level Section"],
            [None, "Subsection"],
            [None, None, "Account A", 100.50],
        ],
    })
    text = extract_xlsx_text(path)
    lines = text.splitlines()
    # "Top Level" has no leading spaces
    assert any(line == "Top Level Section" for line in lines)
    # "Subsection" has 2 spaces
    assert any(line == "  Subsection" for line in lines)
    # "Account A" has 4 spaces
    assert any(line.startswith("    Account A\t") for line in lines)


def test_extract_preserves_cents_precision(tmp_path):
    path = _make_xlsx(tmp_path, {
        "Sheet1": [
            ["Balance", 195243.33, 57907.83, 1234567.89],
        ],
    })
    text = extract_xlsx_text(path)
    # Each dollar value must round-trip without %g truncation
    assert "195243.33" in text
    assert "57907.83" in text
    assert "1234567.89" in text


def test_extract_integer_floats_display_without_trailing_zero(tmp_path):
    path = _make_xlsx(tmp_path, {
        "Sheet1": [["ZeroBalance", 0.0, 2454.0]],
    })
    text = extract_xlsx_text(path)
    # Int-valued floats render as integers (no ".0")
    assert "\t0\t2454" in text


def test_detector_classifies_qb_bs_via_xlsx(tmp_path):
    # Mini QB-BS shape — real QB headers
    path = _make_xlsx(tmp_path, {
        "Sheet1": [
            [None, None, None, None, None, "Dec 31, 25", None, "Dec 31, 24", None, "$ Change"],
            ["ASSETS"],
            [None, "Current Assets"],
            [None, None, "Cash", None, None, 100, None, 100, None, 0],
            [None, "Total Current Assets", None, None, None, 100, None, 100, None, 0],
            ["TOTAL ASSETS", None, None, None, None, 100, None, 100, None, 0],
            ["LIABILITIES & EQUITY"],
            ["TOTAL LIABILITIES & EQUITY", None, None, None, None, 100, None, 100, None, 0],
        ],
    })
    text = extract_xlsx_text(path)
    assert detect_financial_document_type(text) == "BALANCE_SHEET"


def test_detector_classifies_qb_pl_as_comparison(tmp_path):
    # QB P&L with 2 periods + $ Change should be INCOME_STATEMENT_COMPARISON,
    # NOT base INCOME_STATEMENT — the specialization rule should fire.
    path = _make_xlsx(tmp_path, {
        "Sheet1": [
            [None, None, None, None, None, None, "Jan - Dec 25", None, "Jan - Dec 24", None, "$ Change"],
            [None, "Ordinary Income/Expense"],
            [None, None, None, "Income"],
            [None, None, None, None, "Services Income", None, 0, None, 3117100.77, None, -3117100.77],
            [None, None, None, "Total Income", None, None, 0, None, 3117100.77, None, -3117100.77],
            [None, None, None, "Cost of Goods Sold"],
            [None, None, None, "Total COGS", None, None, 0, None, 2096282.20, None, -2096282.20],
            [None, None, "Gross Profit", None, None, None, 0, None, 1024726.09, None, -1024726.09],
            [None, None, None, "Expense"],
            [None, None, None, "Total Expense", None, None, 0, None, 751584.36, None, -751584.36],
            [None, "Net Ordinary Income", None, None, None, None, 0, None, 273141.73, None, -273141.73],
        ],
    })
    text = extract_xlsx_text(path)
    assert detect_financial_document_type(text) == "INCOME_STATEMENT_COMPARISON"


def test_detector_classifies_qb_gl_via_column_header(tmp_path):
    # QB GL with its distinctive column header row
    path = _make_xlsx(tmp_path, {
        "Sheet1": [
            [None, None, None, None, None, "Type", None, "Date", None, "Num", None, "Adj",
             None, "Name", None, "Memo", None, "Split", None, "Debit", None, "Credit", None, "Balance"],
            [None, "Bank of America", None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None, None, None, None, None, None, None, 18167.14],
        ],
    })
    text = extract_xlsx_text(path)
    assert detect_financial_document_type(text) == "GENERAL_LEDGER"


# ---------------------------------------------------------------------------
# Real-fixture integration tests (guarded by skip_if_no_fixture)
# ---------------------------------------------------------------------------


@skip_if_no_fixture("qb_balance_sheet")
def test_qb_balance_sheet_xlsx_detects_as_balance_sheet():
    text = extract_xlsx_text(FIXTURE_REGISTRY["qb_balance_sheet"])
    assert "Total Assets" in text or "TOTAL ASSETS" in text
    assert detect_financial_document_type(text) == "BALANCE_SHEET"


@skip_if_no_fixture("qb_profit_loss")
def test_qb_profit_loss_xlsx_detects_as_comparison():
    text = extract_xlsx_text(FIXTURE_REGISTRY["qb_profit_loss"])
    assert "Gross Profit" in text
    assert detect_financial_document_type(text) == "INCOME_STATEMENT_COMPARISON"


@skip_if_no_fixture("qb_general_ledger")
def test_qb_general_ledger_xlsx_detects_as_gl():
    text = extract_xlsx_text(FIXTURE_REGISTRY["qb_general_ledger"])
    # GL column header
    assert "Debit" in text and "Credit" in text and "Balance" in text
    assert detect_financial_document_type(text) == "GENERAL_LEDGER"

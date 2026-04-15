from __future__ import annotations

from pathlib import Path
from uuid import uuid4

from openpyxl import load_workbook

from app.chunking import build_chunks
from app.extractors.base import BaseExtractor
from app.schemas import (
    DocumentMetadata,
    ExtractionMethod,
    ExtractionPayload,
    ExtractionWarning,
    TextSegment,
)


class XlsxExtractor(BaseExtractor):
    name = "openpyxl"

    def supports(self, filename: str, mime_type: str) -> bool:
        return filename.lower().endswith(".xlsx") or mime_type in {
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "application/vnd.ms-excel",
        }

    def extract(
        self,
        file_path: Path,
        filename: str,
        mime_type: str,
        *,
        ocr_strategy: str = "auto",
        ocr_backend: str = "auto",
    ) -> ExtractionPayload:
        wb = load_workbook(file_path, data_only=True)
        document_id = str(uuid4())
        segments: list[TextSegment] = []
        raw_parts: list[str] = []

        for sheet_idx, sheet in enumerate(wb.worksheets, start=1):
            rows_as_text: list[str] = []
            row_segments: list[TextSegment] = []
            header_values: list[str] | None = None
            for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                values = [str(v).strip() for v in row if v is not None and str(v).strip()]
                if not values:
                    continue
                if header_values is None:
                    header_values = values
                row_text = " | ".join(values)
                rows_as_text.append(row_text)
                metadata = {
                    "sheet_name": sheet.title,
                    "sheet_index": sheet_idx,
                    "row_index": row_idx,
                    "cell_count": len(values),
                    "is_header": row_idx == 1,
                }
                if header_values and row_idx > 1 and len(header_values) == len(values):
                    metadata["header_values"] = header_values
                    metadata["row_mapping"] = dict(zip(header_values, values))
                row_segments.append(
                    TextSegment(
                        type="table",
                        index=row_idx,
                        label=f"{sheet.title}:row-{row_idx}",
                        text=row_text,
                        metadata=metadata,
                    )
                )
            if not rows_as_text:
                continue
            sheet_text = f"Sheet: {sheet.title}\n" + "\n".join(rows_as_text)
            raw_parts.append(sheet_text)
            segments.append(
                TextSegment(
                    type="sheet",
                    index=sheet_idx,
                    label=sheet.title,
                    text=sheet_text,
                    metadata={
                        "sheet_name": sheet.title,
                        "sheet_index": sheet_idx,
                        "row_count": len(row_segments),
                        "header_values": header_values or [],
                    },
                )
            )
            segments.extend(row_segments)

        raw_text = "\n\n".join(raw_parts)
        warnings: list[ExtractionWarning] = []
        status = "success"
        if not segments:
            status = "partial"
            warnings.append(
                ExtractionWarning(
                    code="xlsx_no_text_detected",
                    message="No extractable non-empty cells were detected in this XLSX workbook.",
                )
            )
        return ExtractionPayload(
            document_id=document_id,
            metadata=DocumentMetadata(
                filename=filename,
                mime_type=mime_type,
                source_type="xlsx",
                sheet_names=[ws.title for ws in wb.worksheets],
            ),
            extraction=ExtractionMethod(extractor=self.name, status=status, warnings=warnings),
            raw_text=raw_text,
            segments=segments,
            chunks=build_chunks(document_id, segments),
        )

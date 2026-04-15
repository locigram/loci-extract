"""XLSX/XLS text extraction for QuickBooks Desktop and similar exports.

QB Desktop exports reports as .xlsx where hierarchy is encoded via column
position (the label lives in whichever column corresponds to its depth,
with values in trailing columns). This module reads each sheet via
openpyxl and flattens it into indented text that preserves the visual
structure so the downstream LLM sees a familiar report layout.

Entry point: ``extract_xlsx_text(path) -> str``
"""

from __future__ import annotations

from pathlib import Path

# Sheets to skip — QB Desktop prepends a boilerplate "tips" sheet.
_SKIP_SHEETS = frozenset({
    "QuickBooks Desktop Export Tips",
    "QuickBooks Export Tips",
    "Export Tips",
})


def extract_xlsx_text(path: str | Path) -> str:
    """Read every non-skipped sheet, flatten cells to indented text, join
    with page-break markers so downstream chunking can split per-sheet.

    Returns a single string. Preserves column indentation as leading spaces
    (2 spaces per column index) so the LLM sees nested sections."""
    try:
        import openpyxl
    except ImportError as exc:
        raise RuntimeError(
            "openpyxl is required to read XLSX. Install with `pip install openpyxl`."
        ) from exc

    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    sections: list[str] = []
    for sheet_name in wb.sheetnames:
        if sheet_name in _SKIP_SHEETS:
            continue
        ws = wb[sheet_name]
        lines = _sheet_to_lines(ws)
        if not lines:
            continue
        sections.append(f"--- SHEET: {sheet_name} ---")
        sections.extend(lines)

    return "\n".join(sections).rstrip() + "\n"


def _sheet_to_lines(ws) -> list[str]:
    """Convert one worksheet to a list of text lines.

    Each row becomes: ``<indent spaces><label><tab><value1><tab><value2>...``.
    Empty rows are skipped. Indent is derived from the column index of the
    first non-empty cell (2 spaces per column) so nested sections read
    naturally."""
    lines: list[str] = []
    for row in ws.iter_rows(values_only=True):
        cells = [(ci, v) for ci, v in enumerate(row) if v is not None and str(v).strip()]
        if not cells:
            continue
        # Hierarchy: leftmost populated column = section depth
        first_ci = cells[0][0]
        first_val = str(cells[0][1]).strip()
        indent = "  " * first_ci

        # Remaining cells are the values for this row. In QB exports those
        # are the period columns (amounts). Join with tabs.
        values = [_fmt_cell(cells[i][1]) for i in range(1, len(cells))]
        if values:
            lines.append(f"{indent}{first_val}\t" + "\t".join(values))
        else:
            # Section header row (no values) — keep it on its own line
            lines.append(f"{indent}{first_val}")
    return lines


def _fmt_cell(v) -> str:
    if v is None:
        return ""
    # Preserve exact financial precision. `%g` rounds after 6 sig figs, so
    # 195243.33 becomes 195243 — a disaster for cent-accurate totals. Use
    # repr for floats (round-trip safe), strip trailing ".0" for integers.
    if isinstance(v, float):
        if v.is_integer():
            return str(int(v))
        return repr(v)  # repr is round-trip safe for floats
    return str(v).strip()


__all__ = ["extract_xlsx_text"]
